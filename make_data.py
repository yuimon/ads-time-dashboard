# -*- coding: utf-8 -*-
"""
曜日×時間帯レポート（Google Ads等）から dashboard/data.js を生成するスクリプト。

使い方:
    python make_data.py <レポートのxlsxまたはcsvのパス> [--sheet シート名] [--out 出力先]

例:
    python make_data.py "../20260520-0709.xlsx"
    python make_data.py "新レポート.csv" --out data.js

必要な列（ヘッダー名で自動検出）:
    時間帯（0〜23の数値） / 曜日（日曜日〜土曜日） / 表示回数 / クリック数 / 費用 / コンバージョン
    ※「コンバージョン」より右の列は、CVアクションの内訳として自動的に取り込みます。

期間（例「2026年5月20日 - 2026年7月9日」）がファイル先頭にあれば自動で読み取り、
曜日ごとの日数（「1日あたり平均」補正用）を計算します。見つからない場合、
補正機能はオフになりますがダッシュボード自体は問題なく動きます。
"""
import argparse
import csv
import datetime
import json
import os
import re
import sys

DAYS = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
REQUIRED = ['時間帯', '曜日', '表示回数', 'クリック数', '費用', 'コンバージョン']
DATE_RE = re.compile(r'(\d{4})年(\d{1,2})月(\d{1,2})日')


def read_rows(path, sheet=None):
    """xlsx/csvを行のリスト（各行はセル値のリスト）として読み込む"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xlsx':
        try:
            import openpyxl
        except ImportError:
            sys.exit('openpyxl が必要です。コマンドプロンプトで pip install openpyxl を実行してください。')
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        names = [sheet] if sheet else wb.sheetnames
        best = None
        for name in names:
            if name not in wb.sheetnames:
                sys.exit(f'シート「{name}」が見つかりません。あるのは: {wb.sheetnames}')
            rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
            if find_header(rows) is not None:
                best = (name, rows)
                break
        if best is None:
            sys.exit('「時間帯」と「曜日」の列を持つシートが見つかりませんでした。')
        print(f'シート「{best[0]}」を使用します')
        return best[1]
    if ext == '.csv':
        for enc in ('utf-8-sig', 'utf-16', 'cp932'):
            try:
                with open(path, newline='', encoding=enc) as f:
                    sample = f.read(4096)
                    f.seek(0)
                    delim = '\t' if sample.count('\t') > sample.count(',') else ','
                    return [row for row in csv.reader(f, delimiter=delim)]
            except (UnicodeDecodeError, UnicodeError):
                continue
        sys.exit('CSVの文字コードを判定できませんでした（UTF-8 / UTF-16 / Shift_JIS を試しました）。')
    sys.exit('対応形式は .xlsx と .csv です。')


def find_header(rows):
    """「時間帯」「曜日」を含むヘッダー行の位置を返す"""
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip() if c is not None else '' for c in row]
        if '時間帯' in cells and '曜日' in cells:
            return i
    return None


def to_num(v):
    """' --' やカンマ付き文字列も数値化。数値にならなければ None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('%', '')
    try:
        return float(s)
    except ValueError:
        return None


def parse_period(rows):
    """先頭行から期間を読み取り、(表示用文字列, 曜日ごとの日数) を返す"""
    for row in rows[:5]:
        for c in row:
            if c is None:
                continue
            m = DATE_RE.findall(str(c))
            if len(m) >= 2:
                d1 = datetime.date(*map(int, m[0]))
                d2 = datetime.date(*map(int, m[1]))
                if d1 > d2:
                    d1, d2 = d2, d1
                counts = {d: 0 for d in DAYS}
                cur = d1
                while cur <= d2:
                    counts[DAYS[cur.weekday()]] += 1
                    cur += datetime.timedelta(days=1)
                label = f'{d1.year}年{d1.month}月{d1.day}日 〜 {d2.year}年{d2.month}月{d2.day}日'
                return label, counts
    return None, None


def main():
    ap = argparse.ArgumentParser(description='曜日×時間帯レポートから data.js を生成')
    ap.add_argument('input', help='レポートファイル（.xlsx / .csv）')
    ap.add_argument('--sheet', help='xlsxのシート名（省略時は自動検出）')
    ap.add_argument('--out', default=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.js'),
                    help='出力先（省略時はこのスクリプトと同じ場所の data.js）')
    args = ap.parse_args()

    rows = read_rows(args.input, args.sheet)
    hi = find_header(rows)
    if hi is None:
        sys.exit('「時間帯」と「曜日」の列を持つヘッダー行が見つかりませんでした。')
    header = [str(c).strip() if c is not None else '' for c in rows[hi]]

    col = {}
    for name in REQUIRED:
        if name not in header:
            sys.exit(f'必要な列「{name}」が見つかりません。ヘッダー: {header}')
        col[name] = header.index(name)
    action_cols = [(i, header[i]) for i in range(col['コンバージョン'] + 1, len(header)) if header[i]]

    period, day_counts = parse_period(rows)
    if period is None:
        print('注意: 期間の記載が見つからなかったため、「1日あたり平均」補正は無効になります。')
        period = os.path.basename(args.input)

    data, skipped = [], 0
    for row in rows[hi + 1:]:
        if len(row) < len(REQUIRED):
            continue
        hour = to_num(row[col['時間帯']])
        day = str(row[col['曜日']]).strip() if row[col['曜日']] is not None else ''
        if hour is None or not (0 <= hour <= 23) or day not in DAYS:
            skipped += 1
            continue
        data.append({
            'hour': int(hour),
            'day': day,
            'imp': to_num(row[col['表示回数']]) or 0,
            'clicks': to_num(row[col['クリック数']]) or 0,
            'cost': to_num(row[col['費用']]) or 0,
            'conv': to_num(row[col['コンバージョン']]) or 0,
            'actions': {name: (to_num(row[i]) or 0) for i, name in action_cols},
        })

    if not data:
        sys.exit('有効なデータ行が0件でした。ファイルの形式を確認してください。')

    payload = {'period': period, 'dayCounts': day_counts,
               'source': os.path.basename(args.input), 'rows': data}
    with open(args.out, 'w', encoding='utf-8') as f:
        f.write('const RAW = ' + json.dumps(payload, ensure_ascii=False) + ';\n')

    imp = sum(r['imp'] for r in data)
    clk = sum(r['clicks'] for r in data)
    cost = sum(r['cost'] for r in data)
    conv = sum(r['conv'] for r in data)
    print(f'書き出し完了: {args.out}')
    print(f'  期間: {period}')
    print(f'  データ行: {len(data)}件（スキップ: {skipped}件）')
    print(f'  合計: 表示回数 {imp:,.0f} / クリック {clk:,.0f} / 費用 {cost:,.0f} / CV {conv:,.2f}')
    if len(data) != 168:
        print(f'  注意: 曜日×時間帯の完全なデータは168行ですが、{len(data)}行でした。欠けがないか確認してください。')


if __name__ == '__main__':
    main()
